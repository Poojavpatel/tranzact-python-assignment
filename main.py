import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

path = 'data.xlsx'
wb=load_workbook(path)

lastVisitedLevel = {
  '.1' : False,
  '..2' : False,
  '..3' : False
}
offset = 2

# Function to create a sheet
def createSheet(sheet):
  if(sheet not in wb.sheetnames):
    wb.create_sheet(sheet)
  currentsheet = wb.get_sheet_by_name(sheet)
  currentsheet.append(('Finished Good List','','',''))
  currentsheet.append(('#', 'Item Description', 'Quantity', 'Unit'))
  currentsheet.append(('1', sheet, '1', 'Pc'))
  currentsheet.append(('End of FG', '', '', ''))
  currentsheet.append(('Raw Material List', '', '', ''))
  currentsheet.append(('#', 'Item Description', 'Quantity', 'Unit'))
  wb.save(path)
  return currentsheet

def completeSheet(sheet):
  if(sheet == 'Source'):
    return False
  currentsheet = wb.get_sheet_by_name(sheet)
  currentsheet.append(('End of RM', '', '', ''))
  wb.save(path)
  return currentsheet

def parentCell(level, item):
  if(level == '.1'):
    return item
  if(level == '..2'):
    return lastVisitedLevel['.1']
  if(level == '..3'):
    return lastVisitedLevel['..2']
  return None

sheet=wb.active

for i in range(2,sheet.max_row):
  if(sheet.cell(row=i,column=offset).value is None):
    break
  levelCell=sheet.cell(row=i,column=offset)
  materialCell = sheet.cell(row=i,column=offset+1)
  lastVisitedLevel[levelCell.value] = materialCell.value
  itemCell = sheet.cell(row=i,column=offset-1)
  # figure out parent sheet for multi level BOM
  parent=parentCell(levelCell.value, itemCell.value)
  quantityCell=sheet.cell(row=i,column=offset+2)
  unitCell=sheet.cell(row=i,column=offset+3)
  # If parent sheet does not exist, create new and add values
  if(parent not in wb.sheetnames):
    createSheet(parent)
  currentsheet = wb.get_sheet_by_name(parent)
  currentsheet.append((1,materialCell.value,quantityCell.value,unitCell.value))
  wb.save(path)

# Closing all sheets created except the source sheet
for sheet in wb.sheetnames:
  completeSheet(sheet)